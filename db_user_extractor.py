"""
OS domain user extractor from OS script and write it into Excel file
"""
import os,sys
import re
import openpyxl

pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)
filepath = pathname

sheetName= "DB.xlsx"

wb = openpyxl.load_workbook(sheetName)
sheet = wb.get_active_sheet()
current_ip = 1
startWith = "ORA_U11g_05_STARTS"
endWith = "ORA_U11g_05_ENDS"

while sheet.cell(row=1,column=current_ip).value is not None:
	#print(type(sheet.cell(row=1,column=current_ip).value))
	current_ip += 1

#list down all the files
for foldername in os.listdir(filepath):
	print("Processing : "+foldername)
	ip_with_hostname_finder = re.compile('''
			(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(foldername)
	#print(ip_with_hostname.groups())
	#hostname = ip_with_hostname.group(1)
	ip = ip_with_hostname.group(1)
	#print("################################\n"+str(hostname)+" : "+str(ip))
	# AIX_06 finder
	#if ip is None:
	#	ip = hostname
	db_file = os.listdir(os.path.join(filepath,foldername))
	#print(os.path.join(filepath,foldername,db_file[0]))
	script_file = open(os.path.join(filepath,foldername,'Oracle11g_Unix.txt'),'r')
	script = script_file.read() 
	#print(script)
	# required string extraction
	extracted_string = script[ script.find(startWith): (script.find(startWith)+script.find(endWith))]
	"""extractor_string_finder = re.compile('''
		ORA_U11g_05_STARTS?(\s+)(.+\s+)*ORA_U11g_05_ENDS?
	''',re.X)
	extracted_string = extractor_string_finder.search(script)
	"""
	username_finder = re.compile('''
		(.+)\s*\|(EXPIRED|LOCKED|OPEN)
	''',re.X)
	
	#print(extracted_string)

	usernames = username_finder.findall(extracted_string)
	#print(usernames)
	# for column
	sheet.cell(row=1,column=current_ip).value = foldername
	for users_index in range(1,len(usernames)):
		sheet.cell(row=users_index+1,column=current_ip).value = usernames[users_index][0]
	current_ip += 1
	
wb.save(sheetName)