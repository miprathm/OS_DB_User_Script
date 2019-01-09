"""
OS domain user extractor from OS script and write it into Excel file
"""
import os
import re
import openpyxl
import sys
#filepath = os.path.join(os.getcwd(),"DC")
pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)
filepath = pathname

wb = openpyxl.load_workbook("services.xlsx")
sheet = wb.get_active_sheet()
current_ip = 1

while sheet.cell(row=1,column=current_ip).value is not None:
	#print(type(sheet.cell(row=1,column=current_ip).value))
	current_ip += 1


#list down all the files
for filename in os.listdir(filepath):
	print("Processing : "+filename)
	"""ip_with_hostname_finder = re.compile('''
			(.*)\_audit(\_(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}))?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(filename)
	print(ip_with_hostname.groups())
	hostname = ip_with_hostname.group(1)
	ip = ip_with_hostname.group(3)
	"""
	ip = filename
	hostname = ip
	print("################################\n"+str(hostname)+" : "+str(ip))
	# AIX_06 finder
	if ip is None:
		ip = hostname
	script_file = open(os.path.join(filepath,filename),'r')
	script = script_file.read() 
	aix_section_finder = re.compile('''
			AIX\_26((\s(.*))*?)\s\-+
	''',re.X)
	aix_section_in_script = aix_section_finder.search(script)
	#username_finder
	#print(aix_section_in_script.group(1))
	service_finder = re.compile('''
		(\W?\w*)(\s|\:).*\n
	''',re.X)
	services = service_finder.findall(aix_section_in_script.group(1))
	#print(services)
	# for column
	sheet.cell(row=1,column=current_ip).value = ip
	# logic to remove blanks
	all_service = []
	for index in range(0,len(services)):
		#print("61 "+services[index][0])
		if services[index][0] is not '' :
			temp = (services[index][0]).replace('\n','')
			all_service.append(temp) 
	print(all_service)		
	
	for users_index in range(1,len(all_service)):
		sheet.cell(row=(users_index+1),column=current_ip).value = all_service[users_index]
	current_ip += 1
wb.save("services.xlsx")