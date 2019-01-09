"""
OS domain user extractor from OS script and write it into Excel file
"""
import os
import re
import openpyxl
import sys
#filepath = os.path.join(os.getcwd(),"DC")
pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)
filepath = pathname

wb = openpyxl.load_workbook("OS.xlsx")
sheet = wb.get_active_sheet()
current_ip = 1

while sheet.cell(row=1,column=current_ip).value is not None:
	#print(type(sheet.cell(row=1,column=current_ip).value))
	current_ip += 1


#list down all the files
for filename in os.listdir(filepath):
	print("Processing : "+filename)
	"""ip_with_hostname_finder = re.compile('''
			(.*)\_audit(\_(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}))?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(filename)
	print(ip_with_hostname.groups())
	hostname = ip_with_hostname.group(1)
	ip = ip_with_hostname.group(3)
	"""
	ip = filename
	hostname = ip
	print("################################\n"+str(hostname)+" : "+str(ip))
	# AIX_06 finder
	if ip is None:
		ip = hostname
	script_file = open(os.path.join(filepath,filename),'r')
	script = script_file.read() 
	aix_06_finder = re.compile('''
			AIX\_02((\s(.*))*?)\s\-+
	''',re.X)
	aix_06_in_script = aix_06_finder.search(script)
	#username_finder
	#print(aix_06_in_script.group(1))
	username_finder = re.compile('''
	(.*)\s*
	''',re.X)
	usernames = username_finder.findall(aix_06_in_script.group(1))
	#print(usernames)
	# for column
	sheet.cell(row=1,column=current_ip).value = ip
	for users_index in range(2,len(usernames)):
		sheet.cell(row=users_index,column=current_ip).value = usernames[users_index]
	current_ip += 1
wb.save("OS.xlsx")