"""
OS domain user extractor from OS script and write it into Excel file
"""
import os,sys
import re
import openpyxl


pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)
filepath = pathname

sheetName= "policy.xlsx"

wb = openpyxl.load_workbook(sheetName)
sheet = wb.get_active_sheet()
current_ip = 1
startWith = "ORA_U11g_08_STARTS"
endWith = "ORA_U11g_08_ENDS"

while sheet.cell(row=1,column=current_ip).value is not None:
	#print(type(sheet.cell(row=1,column=current_ip).value))
	current_ip += 1

#list down all the files
for foldername in os.listdir(filepath):
	print("Processing : "+foldername)
	"""ip_with_hostname_finder = re.compile('''
			(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(foldername)
	ip = ip_with_hostname.group(1)
	
	print("################################\n"+str(ip))
	"""
	try:	
		db_file = os.listdir(os.path.join(filepath,foldername))
		script_file = open(os.path.join(filepath,foldername,'Oracle11g_Unix.txt'),'r')
		#print(script_file)
		script = script_file.read() 
		#print(script)
		# required string extraction
		extracted_string = script[ script.find(startWith): (script.find(startWith)+script.find(endWith))]
		if extracted_string is not None:
			failed_login_attempt_finder = re.compile('''
				(\w*)\s*\|FAILED_LOGIN_ATTEMPTS\s*\|(\w*)
			''',re.X)
	
			password_life_time_finder = re.compile('''
				(\w*)\s*\|PASSWORD_LIFE_TIME\s*\|(\w*)
			''',re.X)
			
			password_reuse_max_finder = re.compile('''
				(\w*)\s*\|PASSWORD_REUSE_MAX\s*\|(\w*)
			''',re.X)
			
			password_reuse_time_finder = re.compile('''
				(\w*)\s*\|PASSWORD_REUSE_TIME\s*\|(\w*)
			''',re.X)
	
			password_lock_time_finder = re.compile('''
				(\w*)\s*\|PASSWORD_LOCK_TIME\s*\|(\w*)
			''',re.X)
	
			password_grace_time_finder = re.compile('''
				(\w*)\s*\|PASSWORD_GRACE_TIME\s*\|(\w*)
			''',re.X)
	
			password_verify_function_finder = re.compile('''
				(\w*)\s*\|PASSWORD_VERIFY_FUNCTION\s*\|(\w*)
			''',re.X)
			#print(extracted_string)
		failed_login_attempt = failed_login_attempt_finder.findall(extracted_string)
		password_life_time = password_life_time_finder.findall(extracted_string)
		password_reuse_max = password_reuse_max_finder.findall(extracted_string)
		password_reuse_time = password_reuse_time_finder.findall(extracted_string)
		password_lock_time = password_lock_time_finder.findall(extracted_string)
		password_grace_time = password_grace_time_finder.findall(extracted_string)
		password_verify_function = password_verify_function_finder.findall(extracted_string)
		#print(failed_login_attempt)
		# for column
		sheet.cell(row=1,column=current_ip).value = foldername
		rowIndex = 2;
		for users_index in range(0,len(failed_login_attempt)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = failed_login_attempt[users_index][0]+" | FAILED_LOGIN_ATTEMPTS | "+failed_login_attempt[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_life_time)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_life_time[users_index][0]+" | PASSWORD_LIFE_TIME | "+password_life_time[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_reuse_max)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_reuse_max[users_index][0]+" | PASSWORD_REUSE_MAX | "+password_reuse_max[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_reuse_time)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_reuse_time[users_index][0]+" | PASSWORD_REUSE_TIME | "+password_reuse_time[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_lock_time)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_lock_time[users_index][0]+" | PASSWORD_LOCK_TIME | "+password_lock_time[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_grace_time)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_grace_time[users_index][0]+" | PASSWORD_GRACE_TIME | "+password_grace_time[users_index][1]
			rowIndex += 1

		for users_index in range(0,len(password_verify_function)):
			#parameter
			sheet.cell(row=rowIndex,column=current_ip).value = password_verify_function[users_index][0] +" | PASSWORD_VERIFY_FUNCTION | "+password_verify_function[users_index][1]
			rowIndex += 1
			
		#print(users_index)	
		current_ip += 1
	except:
		print("Error")
wb.save(sheetName)