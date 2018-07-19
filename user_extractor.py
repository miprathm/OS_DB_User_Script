import sys, os
import re
import openpyxl

fill_color = openpyxl.styles.colors.Color(rgb='00ffff00')
fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=fill_color)

#Get path of the folder where output script is available
pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)
#print(sys.argv[1:])

# ================================================== functions for OS Script ================================================== 
# extract AIX_06 block from file
# return string block of AIX_06
def fetch_AIX_06_block(file):
	script_file = open(file,'r')
	script = script_file.read() 
	# search for pattern AIX_06
	aix_06_finder = re.compile('''
			AIX\_06((\s(.*))*?)\s\-+
	''',re.X)
	aix_06_in_script = aix_06_finder.search(script)
	return 	aix_06_in_script.group(1)

# Extract all the OS users from the string
# return list of user	
def fetch_os_users(string):
	username_finder = re.compile('''
	(.+)\:.*\:.*\:.*\:.*\:.*\:.*
	''',re.X)
	return username_finder.findall(string)

# ================================================== End of functions for OS Script ================================================== 

#aix06_block = fetch_AIX_06_block(pathname)
#username = fetch_os_users(aix06_block)	

#print(username)

# ================================================== functions for DB Script ================================================== 
# Extract ORA_U11g_05_STARTS block
def fetch_ORA_U11g_05_block(file):
	startWith = "ORA_U11g_05_STARTS"
	endWith = "ORA_U11g_05_ENDS"
	script_file = open(file,'r')
	script = script_file.read() 
	#print(script)
	# required string extraction
	return script[ script.find(startWith): (script.find(startWith)+script.find(endWith))]
	
def fetch_db_user(string):
	username_finder = re.compile('''
		(.+)\s*\|(EXPIRED|LOCKED|OPEN)
	''',re.X)
	username_list = username_finder.findall(string)
	usernames = []
	for index in range(len(username_list)):
		usernames.append(username_list[index][0].strip())
	return usernames
# ================================================== End of functions for DB Script ================================================== 

#ora_11g_block = fetch_ORA_U11g_05_block(pathname)
#usernames = fetch_db_user(ora_block)

#print(usernames)
def get_ip_os_default_name(name):
	ip_with_hostname_finder = re.compile('''
			(.*)\_audit(\_(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}))?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(name)
	#print(ip_with_hostname.groups())
	ip = ''
	try:
		hostname = ip_with_hostname.group(1)
		ip = ip_with_hostname.group(3)
	except:
		ip = name
		print("Name is unrecognized So take ip as ",ip)
	if ip is None:
		ip = name
	return ip;

def get_ip_db_default_name(name):
	ip_with_hostname_finder = re.compile('''
			(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})?
	''',re.X)
	ip_with_hostname = ip_with_hostname_finder.search(name)
	ip = ip_with_hostname.group(1)
	return ip


def extract_and_save_os_user(path):
	# Create xlsx blank file
	wb = openpyxl.Workbook()
	wb.save('OS.xlsx')
	sheet = wb.get_active_sheet()
	current_ip = 1
	print('--Path -> ',path)	
	for dir in os.listdir(path):
		print('----Dir -> ',dir)
		# for both DC/DR
		for file in os.listdir(os.path.join(path,dir)):
			print('------file -> ',file)
			ip = get_ip_os_default_name(file)
			aix_06_block = fetch_AIX_06_block(os.path.join(path,dir,file))
			usernames = fetch_os_users(aix_06_block)
			sheet.cell(row=1,column=current_ip).fill = fill
			sheet.cell(row=1,column=current_ip).value = ip
			for users_index in range(2,len(usernames)):
				sheet.cell(row=users_index,column=current_ip).value = usernames[users_index]
			current_ip += 1
	wb.save("OS.xlsx")		
#extract_and_save_os_user(pathname)

def extract_and_save_db_user(path):
	wb = openpyxl.Workbook()
	wb.save('DB.xlsx')
	sheet = wb.get_active_sheet()
	current_ip = 1
	print('--Path',path)
	for dcdrdir in os.listdir(path):
		print('----DC/DR Dir -> ',dcdrdir)
		for dir in os.listdir(os.path.join(path,dcdrdir)):
			print('------Dir -> ',dir)
			username = []
			ip = dir#get_ip_db_default_name(dir)
			if os.path.isdir(os.path.join(path,dcdrdir,dir)):
				for file in os.listdir(os.path.join(path,dcdrdir,dir)):
					print("-------- File -> "+file)
					db_user_block = fetch_ORA_U11g_05_block(os.path.join(path,dcdrdir,dir,file))
					username.append('')
					username.append(file)
					username.extend(fetch_db_user(db_user_block))
				sheet.cell(row=1,column=current_ip).value = ip
				sheet.cell(row=1,column=current_ip).fill = fill
				for users_index in range(2,len(username)):
					sheet.cell(row=users_index,column=current_ip).value = username[users_index]
				current_ip += 1	
		wb.save("DB.xlsx")	
#extract_and_save_db_user(pathname)

def processData(path):
	# get dir
	# change CWD
	# create output folder
	# go to each folder create folder new folder in output
	# place DB and OS file into it
	# search for DC/DR folder 
	# execute script on it
	os.chdir(path)
	os.chdir('..')
	os.mkdir('output')
	output_folder = os.path.join(os.getcwd(),"output")
	curr_process_folder = ''
	for folderName in os.listdir(path):
		print("=========================== Processing "+folderName+" ===========================")
		os.chdir(output_folder)
		os.mkdir(folderName)
		curr_process_folder = os.path.join(output_folder,folderName)
		# change current working directory
		os.chdir(curr_process_folder)
		for scriptDir in os.listdir(os.path.join(path,folderName)):
			if re.search('os',scriptDir,re.I):
				print('Processing OS script')
				extract_and_save_os_user(os.path.join(path,folderName,scriptDir,"OS Script Output"))
			elif re.search('db',scriptDir,re.I):
				print('Processing DB script')
				extract_and_save_db_user(os.path.join(path,folderName,scriptDir,"DB Script Output"))
processData(pathname)	